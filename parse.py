"""
_*_ coding:utf-8 _*_
_*_ author:iron_huang _*_
_*_ blog:https://www.dvpos.com/ _*_
"""
import openpyxl
import time
import os
import json
import common
from logger import logger
from script import parse_time, split_line_to_slice, get_set_location, get_loc_new_block_time, get_tipset_key, \
    get_block_header, time_compare, reduce_time_str

info_row = 2
before_remote = 0
after_remote = 0
equ_remote = 0
total_diffs = 0
total_diff_sets = 0


class TipSetInfo:
    def __init__(self, year, month, date, time, height, local_tip_sets, local_tip_sets_num, remote_tip_sets,
                 remote_tip_sets_num, diff_sets, diff_sets_num, is_same, is_jump, jump_from):
        self.year = year
        self.month = month
        self.date = date
        self.time = time
        self.height = height
        self.local_tip_sets = local_tip_sets
        self.local_tip_sets_num = local_tip_sets_num
        self.remote_tip_sets = remote_tip_sets
        self.remote_tip_sets_num = remote_tip_sets_num
        self.diff_sets = diff_sets
        self.diff_sets_num = diff_sets_num
        self.is_same = is_same
        self.is_jump = is_jump
        self.jump_from = jump_from

    def record_base_info(self, row, wsb):
        try:
            wsb.cell(row=row, column=1).value = self.year
            wsb.cell(row=row, column=2).value = self.month
            wsb.cell(row=row, column=3).value = self.date
            wsb.cell(row=row, column=4).value = self.time
            wsb.cell(row=row, column=5).value = self.height

            local_str = "\n".join(self.local_tip_sets)
            wsb.cell(row=row, column=6).value = str(local_str)
            wsb.cell(row=row, column=7).value = self.local_tip_sets_num

            remote_str = "\n".join(self.remote_tip_sets)
            wsb.cell(row=row, column=8).value = str(remote_str)
            wsb.cell(row=row, column=9).value = self.remote_tip_sets_num

            wsb.cell(row=row, column=10).value = self.is_same

            diff_str = "\n".join(self.diff_sets)
            wsb.cell(row=row, column=11).value = str(diff_str)
            wsb.cell(row=row, column=12).value = self.diff_sets_num

            wsb.cell(row=row, column=13).value = self.is_jump
            wsb.cell(row=row, column=14).value = self.jump_from
            print(row - 1)
            time.sleep(0.01)
        except Exception as e:
            logger.error(e)

    def record_info_detail(self, wsi, local_daemon_map):
        try:
            global info_row
            for diff in self.diff_sets:
                wsi.cell(row=info_row, column=1).value = self.height
                wsi.cell(row=info_row, column=2).value = diff
                # 判断所在位置
                loc = get_set_location(diff, self.local_tip_sets)
                wsi.cell(row=info_row, column=3).value = loc
                # 本地出块时间
                loc_time = get_loc_new_block_time(diff, local_daemon_map)
                if loc_time == None:
                    wsi.cell(row=info_row, column=4).value = "未记录"
                    # 远端出块时间
                    miner, time_remote = get_block_header(diff)
                    wsi.cell(row=info_row, column=5).value = time_remote
                    # 出块旷工
                    wsi.cell(row=info_row, column=7).value = miner
                    info_row += 1
                else:
                    wsi.cell(row=info_row, column=4).value = loc_time
                    # 远端出块时间
                    miner, time_remote = get_block_header(diff)
                    wsi.cell(row=info_row, column=5).value = time_remote
                    # 出块旷工
                    wsi.cell(row=info_row, column=7).value = miner
                    # 时间比较
                    compare_result = time_compare(loc_time, time_remote)
                    wsi.cell(row=info_row, column=8).value = compare_result
                    # 计算概率和时间差值
                    global before_remote, after_remote, equ_remote
                    d_value = None
                    if compare_result == common.LOC_BEFORE_REMOTE:
                        d_value = reduce_time_str(time_remote, loc_time)
                        before_remote += 1
                    elif compare_result == common.LOC_AFTER_REMOTE:
                        d_value = reduce_time_str(loc_time, time_remote)
                        after_remote += 1
                    elif compare_result == common.LOC_EQU_REMOTE:
                        equ_remote += 1
                    else:
                        pass
                    wsi.cell(row=info_row, column=6).value = d_value
                    info_row += 1
        except Exception as e:
            logger.error(("%d record detail failed,err:%s" % self.height, e))


def init_local_daemon_map():
    daemon_path = os.path.join(common.DOCS_PATH, common.LOCAL_DAEMON_NAME)
    with open(daemon_path, 'r') as f:
        local_daemon_map = {}
        for line in f:
            before = line.split("new block over pubsub")
            key = json.loads(before[1])["cid"]
            value = " ".join((before[0].split(".")[0]).split("T"))
            local_daemon_map[key] = value
    return local_daemon_map


def init_wb_ws(write_path):
    if os.path.exists(write_path):
        wbl = openpyxl.load_workbook(write_path)
        wslb = wbl.create_sheet(common.DATE_TO_RECORD)
        wsli = wbl.create_sheet(common.DATE_TO_RECORD + "_detail")
    else:
        wbl = openpyxl.Workbook()
        wslb = wbl.create_sheet(common.DATE_TO_RECORD)
        wsli = wbl.create_sheet(common.DATE_TO_RECORD + "_detail")
    wslb.cell(row=1, column=1).value = "年份"
    wslb.cell(row=1, column=2).value = "月份"
    wslb.cell(row=1, column=3).value = "日期"
    wslb.cell(row=1, column=4).value = "时间"
    wslb.cell(row=1, column=5).value = "高度"
    wslb.cell(row=1, column=6).value = "本地tipset"
    wslb.cell(row=1, column=7).value = "本地数量"
    wslb.cell(row=1, column=8).value = "远端tipset"
    wslb.cell(row=1, column=9).value = "远端数量"
    wslb.cell(row=1, column=10).value = "是否相同"
    wslb.cell(row=1, column=11).value = "差别"
    wslb.cell(row=1, column=12).value = "差别数量"
    wslb.cell(row=1, column=13).value = "高度是否断档"
    wslb.cell(row=1, column=14).value = "从何处断档"
    wslb.cell(row=1, column=15).value = "不一致率"

    wsli.cell(row=1, column=1).value = "高度"
    wsli.cell(row=1, column=2).value = "差别"
    wsli.cell(row=1, column=3).value = "位置"
    wsli.cell(row=1, column=4).value = "本地出块时间"
    wsli.cell(row=1, column=5).value = "远端出块时间"
    wsli.cell(row=1, column=6).value = "时间差值"
    wsli.cell(row=1, column=7).value = "出块旷工"
    wsli.cell(row=1, column=8).value = "时间比较"
    wsli.cell(row=1, column=9).value = "早于远端"
    wsli.cell(row=1, column=10).value = "晚于远端"
    wsli.cell(row=1, column=11).value = "等于远端"

    return wbl, wslb, wsli


def record_diff_rates(wsb, wsi, total_loc):
    # 不一致高度占总高度数百分比
    height_diff_rate = "%.2f%%" % ((total_diff_sets / total_loc) * 100)
    wsb.cell(row=2, column=15).value = height_diff_rate
    # 早于远端tipset占总diff百分比
    before_remote_rate = "%.2f%%" % ((before_remote / total_diffs) * 100)
    wsi.cell(row=2, column=9).value = before_remote_rate
    # 晚于远端tipset占总diff百分比
    after_remote_rate = "%.2f%%" % ((after_remote / total_diffs) * 100)
    wsi.cell(row=2, column=10).value = after_remote_rate
    # 等于远端tipset占总diff百分比
    equ_remote_rate = "%.2f%%" % ((equ_remote / total_diffs) * 100)
    wsi.cell(row=2, column=11).value = equ_remote_rate


def new_tip_set_info(year, month, date, time, block_height, local_tip_sets,
                     remote_sets, diff_sets, is_same, is_jump, jump_from):
    final = TipSetInfo(year, month, date, time, block_height, local_tip_sets, len(local_tip_sets), remote_sets,
                       len(remote_sets), diff_sets, len(diff_sets), is_same, is_jump, jump_from)
    return final


def start_parse():
    local_daemon_map = init_local_daemon_map()
    log_path = os.path.join(common.DOCS_PATH, common.LOCAL_MINER_NAME)
    wb, wsb, wsi = init_wb_ws(common.RECORDER_LOCATION)
    has_parsed = []
    with open(log_path, "r") as f:
        line_to_slice = split_line_to_slice(f)
        total_loc = len(line_to_slice)
        base_row = 2
        before = 0
        for piece in line_to_slice:
            is_jump = False
            jump_from = None
            block_height = int(piece[1]["height"])
            if before != 0 and block_height - 1 != before:
                is_jump = True
                jump_from = before
                before = block_height
            else:
                before = block_height
            if block_height in has_parsed:
                continue
            else:
                has_parsed.append(block_height)
                is_same = True
                tip_sets = get_tipset_key(block_height - 1)
                # miner的tipset集合
                local_tip_sets = set(piece[1]["tipset"])
                # filcoin的tipset集合
                remote_sets = set()
                for v in tip_sets:
                    remote_sets.add(v["/"])
                difference_set = remote_sets ^ local_tip_sets
                if len(difference_set) != 0:
                    global total_diffs, total_diff_sets
                    total_diffs += len(difference_set)
                    total_diff_sets += 1
                    is_same = False
                else:
                    pass
                print("record height:%d" % block_height)

                year, month, date, time = parse_time(piece[0])

                final = new_tip_set_info(year, month, date, time, block_height, local_tip_sets,
                                         remote_sets, difference_set, is_same, is_jump, jump_from)
                final.record_base_info(base_row, wsb)
                if len(difference_set) != 0:
                    final.record_info_detail(wsi, local_daemon_map)
                else:
                    pass
                base_row += 1
    record_diff_rates(wsb, wsi, total_loc)
    wb.save(common.RECORDER_LOCATION)
    wb.close()
    print("record over!")
