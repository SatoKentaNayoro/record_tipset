"""
_*_ coding:utf-8 _*_
_*_ author:iron_huang _*_
_*_ blog:https://www.dvpos.com/ _*_
"""

import paramiko
import os
import time
import json
import common
import openpyxl
import requests

base_path = os.path.abspath("../")
local_daemon_path = os.path.join(base_path, "docs/daemontodo.txt")
record_path = os.path.join(base_path, "docs/daemontodo.xlsx")


def exec_copy_cmd(scp):
    sftp = paramiko.SFTPClient.from_transport(scp)
    try:
        sftp.get("daemontodo.log", local_daemon_path)
        scp.close()
        return "copy ok"
    except Exception as e:
        return e


def init_scp(pwd):
    scp = paramiko.Transport(common.TARGET_IP, common.PORT)
    scp.connect(username=common.USER_NAME, password=pwd)
    return scp


def start_copy(pwd):
    scp = init_scp(pwd)
    print("start copy new log file")
    msg = exec_copy_cmd(scp)
    print(msg)


def get_block_header(cid):
    headers = {'Accept': 'application/json',
               'Content-Type': 'application/json',
               'Authorization': 'Bearer ' + common.TOKEN}
    params = {
        "jsonrpc": "2.0",
        "method": "Filecoin.ChainGetBlock",
        "id": 1,
        "params": [
            {"/": cid},
        ]
    }
    r = requests.post(common.URL, headers=headers, json=params)
    if r.status_code == 200:
        result = r.json().get("result")
        if result != None:
            return result
    else:
        return None


def init_local_daemon_map():
    with open(local_daemon_path, 'r') as f:
        local_daemon_map = {}
        for line in f:
            before = line.split("new block over pubsub")
            key = json.loads(before[1])["cid"]
            value = " ".join((before[0].split(".")[0]).split("T"))
            local_daemon_map[key] = value
    return local_daemon_map


def record_all_info(local_daemon_map, ws):
    count6, count15, count20, count_over20, total_count = 0, 0, 0, 0, 0
    row = 2
    for k in local_daemon_map:
        total_count += 1
        block_head = get_block_header(k)
        loc_time = local_daemon_map.get(k)
        if block_head == None:
            ws.cell(row, column=2).value = k
            ws.cell(row, column=3).value = loc_time
            print(("record %d") % (row - 1))
            row += 1
        else:
            height, remote_time_stamp, miner = block_head["Height"], block_head["Timestamp"], block_head["Miner"]
            time_array = time.localtime(remote_time_stamp)
            time_remote = time.strftime("%Y-%m-%d %H:%M:%S", time_array)
            loc_time_array = time.strptime(loc_time, "%Y-%m-%d %H:%M:%S")
            loc_stamp = int(time.mktime(loc_time_array))
            cost = loc_stamp - remote_time_stamp
            if cost > 6:
                if 6 < cost <= 10:
                    count6 += 1
                elif 10 < cost <= 15:
                    count15 += 1
                elif 15 < cost <= 20:
                    count20 += 1
                else:
                    count_over20 += 1
                print(("record %d,height: %d") % ((row - 1), height))
                ws.cell(row, column=1).value = height
                ws.cell(row, column=2).value = k
                ws.cell(row, column=3).value = loc_time
                ws.cell(row, column=4).value = time_remote
                ws.cell(row, column=5).value = str(cost) + "s"
                ws.cell(row, column=6).value = miner
                row += 1
            else:
                pass
        time.sleep(0.01)
    ws.cell(2, column=7).value = "%.2f%%" % ((count6 / total_count) * 100)
    ws.cell(2, column=8).value = "%.2f%%" % ((count15 / total_count) * 100)
    ws.cell(2, column=9).value = "%.2f%%" % ((count20 / total_count) * 100)
    ws.cell(2, column=10).value = "%.2f%%" % ((count_over20 / total_count) * 100)


def init_wb_ws(write_path):
    if os.path.exists(write_path):
        wbl = openpyxl.load_workbook(write_path)
        wslb = wbl.create_sheet("all_daemon")
    else:
        wbl = openpyxl.Workbook()
        wslb = wbl.create_sheet("all_daemon")
    wslb.cell(row=1, column=1).value = "高度"
    wslb.cell(row=1, column=2).value = "cid"
    wslb.cell(row=1, column=3).value = "本地时间"
    wslb.cell(row=1, column=4).value = "远端时间"
    # wslb.cell(row=1,column=5).value = ""
    wslb.cell(row=1, column=5).value = "相差时间"
    wslb.cell(row=1, column=6).value = "矿工"
    wslb.cell(row=1, column=7).value = "6s-10s百分比"
    wslb.cell(row=1, column=8).value = "10s-15s百分比"
    wslb.cell(row=1, column=9).value = "15s-20s百分比"
    wslb.cell(row=1, column=10).value = "超过20s百分比"

    return wbl, wslb


if __name__ == '__main__':
    start_copy("abmabm123678")
    local_daemon_map = init_local_daemon_map()
    wb, ws = init_wb_ws(record_path)
    record_all_info(local_daemon_map, ws)
    wb.save(record_path)
    wb.close()
    print("record over !")
